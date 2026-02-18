"""
QA Agent — Unit tests for the exporter module.

Tests Excel generation and Google Sheets export (mocked).
"""

import io
import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))

from exporter import HEADERS, _build_full_name, export_to_excel, export_to_google_sheet


# ═══════════════════════════════════════════════════════════════════
# _build_full_name tests
# ═══════════════════════════════════════════════════════════════════


class TestBuildFullName:
    def test_all_parts(self):
        rec = {"lastname": "Dela Cruz", "firstname": "Juan", "middlename": "Santos", "extension": "Jr."}
        assert _build_full_name(rec) == "Dela Cruz, Juan Santos Jr."

    def test_no_middle_no_ext(self):
        rec = {"lastname": "Reyes", "firstname": "Maria", "middlename": "", "extension": ""}
        assert _build_full_name(rec) == "Reyes, Maria"

    def test_only_lastname(self):
        rec = {"lastname": "Cruz", "firstname": "", "middlename": "", "extension": ""}
        assert _build_full_name(rec) == "Cruz"

    def test_empty(self):
        rec = {"lastname": "", "firstname": "", "middlename": "", "extension": ""}
        assert _build_full_name(rec) == ""


# ═══════════════════════════════════════════════════════════════════
# Excel export tests
# ═══════════════════════════════════════════════════════════════════


SAMPLE_RECORDS = [
    {
        "full_name": "Dela Cruz, Juan Santos",
        "email": "",
        "gender": "M",
        "beneficiary": "Youth",
        "age_range": "15-20",
        "affiliation_type": "University",
        "affiliation_name": "University of Bohol",
    },
    {
        "full_name": "Reyes, Maria Lopez Jr.",
        "email": "maria@test.com",
        "gender": "F",
        "beneficiary": "",
        "age_range": "",
        "affiliation_type": "",
        "affiliation_name": "",
    },
]


class TestExportToExcel:
    def test_returns_bytes_stream(self):
        stream = export_to_excel(SAMPLE_RECORDS)
        assert isinstance(stream, io.BytesIO)
        assert stream.tell() == 0  # cursor at start

    def test_correct_headers(self):
        stream = export_to_excel(SAMPLE_RECORDS)
        wb = load_workbook(stream)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        assert header_row == HEADERS

    def test_correct_data_rows(self):
        stream = export_to_excel(SAMPLE_RECORDS)
        wb = load_workbook(stream)
        ws = wb.active

        # Row 2 = first data row  (Full Name, Email, Gender, Beneficiary, Age Range, Affiliation Type, Affiliation Name)
        assert ws.cell(row=2, column=1).value == "Dela Cruz, Juan Santos"
        assert ws.cell(row=2, column=2).value in ("", None)  # empty email
        assert ws.cell(row=2, column=3).value == "M"
        assert ws.cell(row=2, column=4).value == "Youth"
        assert ws.cell(row=2, column=5).value == "15-20"
        assert ws.cell(row=2, column=6).value == "University"
        assert ws.cell(row=2, column=7).value == "University of Bohol"

        # Row 3 = second data row
        assert ws.cell(row=3, column=1).value == "Reyes, Maria Lopez Jr."
        assert ws.cell(row=3, column=2).value == "maria@test.com"
        assert ws.cell(row=3, column=3).value == "F"

    def test_row_count(self):
        stream = export_to_excel(SAMPLE_RECORDS)
        wb = load_workbook(stream)
        ws = wb.active
        assert ws.max_row == 3  # 1 header + 2 data rows

    def test_empty_records(self):
        stream = export_to_excel([])
        wb = load_workbook(stream)
        ws = wb.active
        assert ws.max_row == 1  # header only

    def test_worksheet_name(self):
        stream = export_to_excel(SAMPLE_RECORDS)
        wb = load_workbook(stream)
        assert wb.active.title == "Attendance"


# ═══════════════════════════════════════════════════════════════════
# Google Sheets export tests (mocked)
# ═══════════════════════════════════════════════════════════════════


class TestExportToGoogleSheet:
    def test_missing_credentials_raises(self, tmp_path):
        fake_path = tmp_path / "nonexistent.json"
        with pytest.raises(FileNotFoundError, match="credentials"):
            export_to_google_sheet(SAMPLE_RECORDS, fake_path)

    @patch("exporter.gspread")
    @patch("exporter.ServiceAccountCredentials")
    def test_creates_sheet_and_writes(self, mock_creds_cls, mock_gspread, tmp_path):
        # Create a fake credentials file
        creds_file = tmp_path / "service_account.json"
        creds_file.write_text('{"type": "service_account"}')

        mock_creds = MagicMock()
        mock_creds_cls.from_service_account_file.return_value = mock_creds

        mock_worksheet = MagicMock()
        mock_spreadsheet = MagicMock()
        mock_spreadsheet.worksheet.return_value = mock_worksheet
        mock_spreadsheet.url = "https://docs.google.com/spreadsheets/d/test"

        mock_gc = MagicMock()
        mock_gc.open.return_value = mock_spreadsheet
        mock_gspread.authorize.return_value = mock_gc

        url = export_to_google_sheet(SAMPLE_RECORDS, creds_file, "Test Sheet", "Data")

        assert url == "https://docs.google.com/spreadsheets/d/test"
        mock_worksheet.clear.assert_called_once()
        mock_worksheet.update.assert_called_once()

        # Verify the data sent to update
        call_args = mock_worksheet.update.call_args
        rows = call_args[0][0]
        assert rows[0] == HEADERS  # first row = headers
        assert len(rows) == 3  # header + 2 records
        assert rows[1][0] == "Dela Cruz, Juan Santos"  # Full Name

    @patch("exporter.gspread")
    @patch("exporter.ServiceAccountCredentials")
    def test_creates_new_spreadsheet_when_not_found(self, mock_creds_cls, mock_gspread, tmp_path):
        import gspread

        creds_file = tmp_path / "service_account.json"
        creds_file.write_text('{"type": "service_account"}')

        mock_creds = MagicMock()
        mock_creds_cls.from_service_account_file.return_value = mock_creds

        mock_worksheet = MagicMock()
        mock_new_spreadsheet = MagicMock()
        mock_new_spreadsheet.worksheet.return_value = mock_worksheet
        mock_new_spreadsheet.url = "https://docs.google.com/spreadsheets/d/new"

        mock_gc = MagicMock()
        mock_gc.open.side_effect = gspread.SpreadsheetNotFound
        mock_gc.create.return_value = mock_new_spreadsheet
        mock_gspread.authorize.return_value = mock_gc
        mock_gspread.SpreadsheetNotFound = gspread.SpreadsheetNotFound

        url = export_to_google_sheet(SAMPLE_RECORDS, creds_file)

        mock_gc.create.assert_called_once_with("Attendance Export")
        assert "new" in url
