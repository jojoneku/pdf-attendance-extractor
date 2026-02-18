"""
Export module for attendance data.

Supports exporting extracted student records to:
  - Excel (.xlsx) via openpyxl
  - Google Sheets via gspread + service account
"""

from __future__ import annotations

import io
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    import gspread
    from google.oauth2.service_account import Credentials as ServiceAccountCredentials
except ImportError:
    gspread = None  # type: ignore[assignment]
    ServiceAccountCredentials = None  # type: ignore[assignment, misc]

HEADERS = ["Full Name", "Email", "Gender", "Beneficiary", "Age Range", "Affiliation Type", "Affiliation Name"]


def _build_full_name(record: dict) -> str:
    """Build 'Lastname, Firstname Middlename Extension' from parts."""
    last = (record.get("lastname") or "").strip()
    first = (record.get("firstname") or "").strip()
    middle = (record.get("middlename") or "").strip()
    ext = (record.get("extension") or "").strip()

    # "DELA CRUZ, JUAN SANTOS JR"
    name_parts = [first]
    if middle:
        name_parts.append(middle)
    if ext:
        name_parts.append(ext)
    after_comma = " ".join(name_parts)

    if last and after_comma:
        return f"{last}, {after_comma}"
    return last or after_comma or ""


def export_to_excel(records: list[dict], filename: str = "attendance_export.xlsx") -> io.BytesIO:
    """
    Generate an Excel workbook from aggregated student records.

    Args:
        records: List of dicts with keys matching HEADERS (lowercase/snake_case).
        filename: Not used for the stream; kept for reference.

    Returns:
        BytesIO stream containing the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # ── Header row styling ──────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # ── Data rows ──────────────────────────────────────────────────
    field_keys = ["full_name", "email", "gender", "beneficiary", "age_range", "affiliation_type", "affiliation_name"]

    for row_idx, record in enumerate(records, start=2):
        for col_idx, key in enumerate(field_keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=record.get(key, ""))

    # ── Auto-fit column widths ─────────────────────────────────────
    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3

    # ── Write to stream ────────────────────────────────────────────
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def export_to_google_sheet(
    records: list[dict],
    credentials_path: str | Path,
    spreadsheet_name: str = "Attendance Export",
    worksheet_name: str = "Sheet1",
) -> str:
    """
    Push aggregated student records to a Google Sheet.

    Args:
        records: List of dicts with student data.
        credentials_path: Path to the Google service account JSON file.
        spreadsheet_name: Name of the spreadsheet to create or open.
        worksheet_name: Name of the worksheet tab.

    Returns:
        URL of the Google Sheet.

    Raises:
        FileNotFoundError: If credentials file is missing.
        Exception: On Google API errors.
    """
    creds_path = Path(credentials_path)
    if not creds_path.exists():
        raise FileNotFoundError(
            f"Google credentials file not found: {creds_path}. "
            "See README.md for setup instructions."
        )

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    credentials = ServiceAccountCredentials.from_service_account_file(str(creds_path), scopes=scopes)
    gc = gspread.authorize(credentials)

    # Try to open existing, or create new
    try:
        spreadsheet = gc.open(spreadsheet_name)
    except gspread.SpreadsheetNotFound:
        spreadsheet = gc.create(spreadsheet_name)

    # Get or create worksheet
    try:
        worksheet = spreadsheet.worksheet(worksheet_name)
    except gspread.WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(title=worksheet_name, rows=len(records) + 1, cols=len(HEADERS))

    # Build rows: header + data
    field_keys = ["full_name", "email", "gender", "beneficiary", "age_range", "affiliation_type", "affiliation_name"]
    rows = [HEADERS]
    for record in records:
        rows.append([record.get(key, "") for key in field_keys])

    # Batch update (single API call)
    worksheet.clear()
    worksheet.update(rows, value_input_option="USER_ENTERED")

    # Format header row bold
    worksheet.format("1:1", {"textFormat": {"bold": True}})

    return spreadsheet.url
